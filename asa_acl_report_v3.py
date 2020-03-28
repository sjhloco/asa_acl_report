#!/usr/bin/env python
import argparse
import os
import re
from os.path import expanduser
from pprint import pprint
from getpass import getpass
from netmiko import Netmiko
from ipaddress import ip_address
from ipaddress import IPv4Network
from datetime import datetime
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, colors, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

############# USER DEFINED VARIABLES #############
directory = expanduser("~")
device = 'ste@10.10.10.1'
report_name = device.split('@')[1] + '_ACLreport_' + date.today().strftime('%Y%m%d')
# Header names and columns widths for the XL sheet
header = {'ACL Name':22, 'Line Number':17, 'Access':10, 'Protocol':12, 'Source Address':19, 'Source Port':16, 'Destination Address':24,
          'Destination Port':20, 'Hit Count':14, 'Date Last Hit':17, 'Time Last Hit':17, 'State':10}

################################## 1. User input collected ##################################
# Optional flags user can enter to customise what is run, if nothing is entered uses the default options
def create_parser():
	parser = argparse.ArgumentParser()
	parser.add_argument('-f', '--filename', nargs='*', help="Name of a files (ACL and optionally acl brief) in 'directory' to run the script against.")
	parser.add_argument('-d', '--device', default=device, help='Username and device (username@device) to run the script against (default: %(default)s)')
	parser.add_argument('-i', '--ip', nargs='*', help='IP addresses or networks to filter on')
	parser.add_argument('-a', '--acl', nargs='*', help='ACL names to filter on')
	parser.add_argument('-l', '--location', default=directory, help='Location to save and run the report (default: %(default)s)')
	parser.add_argument('-n', '--name', default=report_name, help='Name for the report (default: %(default)s)')
	return vars(parser.parse_args())

################################## 2. Validate info fed into the script ##################################
class Validate():
    def __init__(self, args):
        self.filename = args['filename']
        self.device = args['device']
        self.filter_ip = args['ip']
        self.filter_acl = args['acl']
        self.location = args['location']
        self.name = args['name']

# Verify the arguments entered are valid
    def verify_args(self):
        print('Checking the options entered are valid...')
        file_error, file_acl_error, self.asa_all_acls, file_all_acls, acl_error, ip_error = ([] for i in range(6))
        self.against_asa = False         # Decides whether the ASA configuration or file is used for the input

        # FILENAME: Checks that the acl file and optionally the acl brief file exist, creates list of any that dont
        if self.filename != None:
            if not os.path.exists(os.path.join(self.location, self.filename[0])):
                file_error.append(os.path.join(self.location, self.filename[0]))
            if len(self.filename) == 2:
                if not os.path.exists(os.path.join(self.location, self.filename[1])):
                    file_error.append(os.path.join(self.location, self.filename[1]))

            # Normalises file to remove non-ace lines
            if len(file_error) == 0:
                with open(os.path.join(self.location, self.filename[0])) as var:
                    self.file_acl = var.read().splitlines()
                    for ace in list(self.file_acl):
                        if (len(ace) == 0) or ('show' in ace) or ('access-list' not in ace) or ('elements' in ace) or ('cached' in ace) or ('remark' in ace):
                            self.file_acl.remove(ace)
                # Checks for hitcnt in ACL file (if missing means is the show run access-list)
                for ace in self.file_acl:
                    if 'hitcnt' not in ace:
                        file_acl_error.append(ace)
                # ACL: Makes sure that the ACL names entered exist in the file, creates list of any that do not
                if self.filter_acl != None:
                    for ace in self.file_acl:
                        y = ace.lstrip()
                        y = y.split(' ')
                        file_all_acls.append(y[1])
                    # Finds any specified ACLs not in the file
                    acl_error = set(self.filter_acl) - set(file_all_acls)

            # NAME: If the file already exists user is asked whether they wish to overwrite it
            file_exist = "yes"
            while file_exist == "yes":
                if os.path.exists(os.path.join(self.location, self.name + ".xlsx")):
                    print("The output file already exist, do you want to overwrite it?")
                    answer = input('y or n: ')
                    if answer == 'n':
                        self.name = input("Please enter a new name for the output file: ")
                    elif answer == 'y':
                        file_exist = "no"
                    else:
                        print("!!! Error - The only acceptable options are 'y' or 'n' !!!")
                else:
                    file_exist = "no"

        # DEVICE: Validate device is reachable and the username and password are correct
        elif self.filename == None:
            self.against_asa = True
            while True:
                try:
                    password = getpass('Enter the ASA password: ')
                    self.net_conn = Netmiko(host=self.device.split('@')[1], username=self.device.split('@')[0], password=password, device_type='cisco_asa')
                    self.net_conn.find_prompt()         # Expects to receive prompt back from the ASA
                    break
                except Exception as e:                  # If login fails loops to begining displaying the error message
                    print(e)

            # ACL: Makes sure that the ACL names entered exist on the ASA, creates list of any that do not
            asa_acl = self.net_conn.send_command('show run access-group')
            vpn_acl = self.net_conn.send_command('show run | in split-tunnel-network-list')
            for ace in asa_acl.splitlines():
                self.asa_all_acls.append(ace.split(' ')[1])
            for ace in vpn_acl.splitlines():
                self.asa_all_acls.append(ace.split('value ')[1])
            # Converts to a 'set' to remove duplicates, then finds any specified ACLs not on ASA
            if self.filter_acl != None:
                acl_error = set(self.filter_acl) - set(self.asa_all_acls)

        # IP: Makes sure the IP address are valid, if creates a list of non-valid IPs
        if self.filter_ip != None:
            for ip_addr in self.filter_ip:
                try:
                    ip_address(ip_addr)
                except ValueError as errorCode:
                    ip_error.append(str(errorCode))

        # ERROR RESULTS: If any of the above generate errors (lists not empty) interact with user and kill the script
        if len(file_error) != 0:
             print('!!! ERROR - Invalid filenames entered, {} do not exist !!!'.format(file_error))
        if len(file_acl_error) !=0:
            print('!!! ERROR - No hitcnt in {}, make sure is output of "show RUN access-list" and NOT "show access-list" !!!'.format(self.filename[0]))
        if len(acl_error) != 0:
            print("!!! ERROR - Invalid ACL names entered, {} are either not in the file or not ACLs on the ASA !!!".format(acl_error))
        if len(ip_error) != 0:
            print("!!!ERROR - The following IP address are not valid !!!")
            for x in ip_error:
                print(x)
        if len(file_error) != 0 or len(file_acl_error) !=0 or len(acl_error) != 0 or len(ip_error) != 0:
            exit()

        return [self.filename, self.device, self.filter_ip, self.filter_acl, self.location, self.name]              # Used for unit testing

################################## 3. Get and filter ACLs based on IP and name ##################################
    def get_filter_acl(self):
        acl_brief = []
        acl, acl_brief_temp = ('' for i in range(2))
    	# Connects to ASA and gets ACL and ACL brief based on the filters
        if self.against_asa == True:
            print('Gathering ACL info from the ASA...')
    	    # Gathers ACL brief output for all ACLs on the ASA
            for ace in set(self.asa_all_acls):
                acl_brief_temp = acl_brief_temp + self.net_conn.send_command('show access-list {} brief'.format(ace))
            #ALL FILTER: Get all ACE entries from all ACLs
            if self.filter_ip == None and self.filter_acl == None:
                acl = self.net_conn.send_command('show access-list | ex elements|cached|alert-interval|remark')
            #ALL FILTER: To get specific ACE entries from the specified ACLs
            elif self.filter_ip != None and self.filter_acl != None:
                self.filter_ip = '|'.join(self.filter_ip)
                for ace in self.filter_acl:
                    acl = acl + self.net_conn.send_command('show access-list {} | in {}'.format(ace, self.filter_ip))
            #ACL FILTER: To get all ACE entries from the specified ACLs names
            elif self.filter_acl != None:
                for ace in self.filter_acl:
                    acl = acl + self.net_conn.send_command('show access-list {}'.format(ace))
            #IP FILTER: To get specific ACE entries (based on ip) from all ACLs
            elif self.filter_ip != None:
                self.filter_ip = '|'.join(self.filter_ip)
                acl = self.net_conn.send_command('show access-list | in {}'.format(self.filter_ip))
            self.net_conn.disconnect()

        # Searches through the ACL file to filter only specified IPs and/or ACLs
        elif self.against_asa == False:
            print('Gathering ACL info from the File...')
            acl, acl_temp = ([] for i in range(2))
            #ALL FILTER: To get all ACL entries from all ACLs
            if self.filter_ip == None and self.filter_acl == None:
                acl = self.file_acl
            #ALL FILTER: To get specific ACE entries from the specified ACLs
            elif self.filter_ip != None and self.filter_acl != None:
                for filter_ace in self.filter_acl:
                    for file_ace in self.file_acl:
                        if filter_ace in file_ace:
                            acl_temp.append(file_ace)
                for filter_ip in self.filter_ip:
                    for ace in acl_temp:
                        if filter_ip in ace:
                            acl.append(ace)
            #ACL FILTER: To get all ACE entries from the specified ACL names
            elif self.filter_acl != None:
                for filter_ace in self.filter_acl:
                    for file_ace in self.file_acl:
                        if filter_ace in file_ace:
                            acl.append(file_ace)
    		#IP FILTER: To get specific ACE entries (based on ip) from all ACLs
            elif self.filter_ip != None:
                for filter_ip in self.filter_ip:
                    for file_ace in self.file_acl:
                        if filter_ip in file_ace:
                            acl.append(file_ace)

            # Converts the ACL file back from a list into a string and runs next function
            acl = '\n'.join(acl)

            # If the ACL brief file exists it is loaded into a variable
            if len(self.filename) == 2:
                with open(os.path.join(self.location, self.filename[1])) as f:
                    acl_brief_temp = f.read()

        # If run against an ASA or has an ACL brief file it creates new list of all lines that have a timestamp (matching 8 characters, space, 8 characters)
        if len(acl_brief_temp) != 0:
            for x in acl_brief_temp.splitlines():
                if re.match(r"^\S{8}\s\S{8}\s", x):
                    acl_brief.append(x)
        return [acl, acl_brief]

################################## 4. Sanitize the data - Create structured data ##################################
class Format_data():
    def __init__(self, acl_data):
        self.acl = acl_data[0]
        self.acl_brief = acl_data[1]

    def format_acl(self):
        print('Formatting the ACL data...')
        # 4a. Pad out 'any' and remove 'hitcnt' text by replacing fields. Is all done whilst file is one big string
        self.acl = self.acl.replace('any4', 'any').replace('any', 'any any1')    # Swap any4 from packet captures then padout any
        for elem in ['(hitcnt=', ')']:
            self.acl = self.acl.replace(elem, '')       # Removeing hitcnt to just leave the value

        # 4b. Clean up ACL by removing remarks and informational data (from ASA output, file was already done), produces a list
        self.acl = self.acl.splitlines()
        for x in self.acl:              # Have to use the while loop as .remove() only removes first match from the list
            while ('elements' in x) or ('cached' in x) or ('alert-interval' in x) or ('remark' in x) or ('standard' in x) or (len(x) == 0):
                for x in self.acl:
                    if ('elements' in x) or ('cached' in x) or ('alert-interval' in x) or ('remark' in x) or ('standard' in x) or (len(x) == 0):
                        self.acl.remove(x)

        # 4c. Remove unneeded fields, and lines with object to leave only the actual data
        acl_temp1, acl_temp2, acl_temp3, acl_temp4, acl_temp5 = ([] for i in range(5))
        for ace in self.acl:
            if 'object' not in ace:     	    # Remove all lines with object
                ace_1 = ace.strip().split(' ')  # Removes starting/ trailing whitespaces before splitting at any other whitespaces
                for field in [0, 1, 2]:         # Deletes first 3 fields (access-list, line and extended)
                    del ace_1[field]
                acl_temp1.append(ace_1)

        # 4d. Normalise source ports by joining range, removing eq and padding out if their is no source port
        for ace in acl_temp1:
            if ace[6] == 'range':             	# If has a source range of ports replace "range" with "start-end" port numbers
                start = ace.pop(7)
                end = ace.pop(7)
                ace[6] = start + '-' + end
            elif ace[6] == 'eq':              	# If has a single source port delete eq
                del ace[6]
            else:           	# If it is ICMP (cant have src_port) or has no source port padout with 'any_port'
                (ace.insert(6, 'any_port'))
            acl_temp2.append(ace)

        # 4e. Normalise destination ports by joining range, removing eq and padding out if no source port
        for ace in acl_temp2:
            if 'range' in ace:
                start = ace.pop(10)
                end = ace.pop(10)
                ace[9] = start + '-' + end
            elif 'eq' in ace:
                del ace[9]
            elif 'icmp' not in ace:
                (ace.insert(9, 'any_port'))
            elif ace[9].isdigit() or ace[9] == 'log':
                (ace.insert(9, 'any_port'))
            acl_temp3.append(ace)

        # 4f. If ACL is logging removes log and all fields up to the hitcnt and hash
        for ace in acl_temp3:
            if 'log' in ace:
                del ace[10:-2]
            acl_temp4.append(ace)

       # 4g. If ACL is inatcive removes extra columns and adds 'inactive to last column
        for ace in acl_temp4:
            if 'inactive' in ace:
                del ace[-2]
                ace.append(ace.pop(ace.index('inactive')))
            acl_temp5.append(ace)
        # All lines now [name, num, permit/deny, protocol, src_ip, src_mask, src_port, dst_ip, dst_mask, dst_port, hitcnt, hash, inactice]

        # 4h. For all host entries delete host and add /32 subnet mask after the IP
        acl = []
        for ace in acl_temp5:
            if ace[4] == 'host':						# if src_ip is /32
                del ace[4]
                (ace.insert(5, '255.255.255.255'))
            if ace[7] == 'host':						# if dst_ip is /32
                del ace[7]
                (ace.insert(8, '255.255.255.255'))
            acl.append(ace)

        # 4i. Convert subnet mask to prefix and padout old mask with 'any1'. If interface name used in ace do same for that
        for ace in acl:
            if ace[4] == 'interface':										# If interface is used in the ace change name for 'any1'
                ace[5] = 'any1'
            elif ace[4] != 'any':										        # As long as is not 'any'
                src_pfx = IPv4Network((ace[4], ace[5])).with_prefixlen	    # Add prefix to the src_IP
                ace[4] = src_pfx
                ace[5] = 'any1'										        # Change old mask to 'any1'
            if ace[7] == 'interface':
                ace[5] = 'any1'
            elif ace[7] != 'any':
                dst_pfx = IPv4Network((ace[7], ace[8])).with_prefixlen		# Add prefix to the src_IP
                ace[7] = dst_pfx
                ace[8] = 'any1'
            del ace[5]       											    # Delete the 'any1' padding that was used earlier
            del ace[7]
        # Format will now be [name, num, permit/deny, protocol, src_ip/pfx, src_port, dst_ip/pfx, dst_port, hitcnt, hash]
        return(acl)

################################## 5. Add ACE last hit timestamp ##################################

    # Adds date and time for last hit for any ACEs with a hitcnt. This info is got as unix hash from last element in show access-list <name> brief
    def lasthit_time(self, acl):
        # If ACE hash matches acl_brief hash the last hit date and time will be added
        for ace in acl:              						# loop through ACL
            ace.insert(10, '')                              # Insert blank column to be used by the time
            for hashes in self.acl_brief:     				# Loop through acl_brief
                if hashes.split(' ')[0] in ace[9]:    		# If acl_brief hash matches ace hash
                    unix_time = hashes.split(' ')[-1]
                    # Convert last element to decimal and get human-readable time from the hex time
                    human_time = datetime.fromtimestamp(int(unix_time, 16)).strftime('%Y-%m-%d %H:%M:%S').split(' ')
                    ace[9] = human_time[0]					# Replaces hash with the date
                    ace[10] = human_time[1]			    # Adds time to the new blank column that was added
        # If ACE does not have a matching hash entry hash is stripped out
        for lines in acl:
            if len(lines[10]) == 0:
                lines[9] = None
        return(acl)

################################## 6. Build XL worksheet ##################################

def create_xls(args, final_acl):
    print('Creating the spreadsheet...')
    filename = os.path.join(args[4], args[5] + ".xlsx")
    # Create workbook with the one defaut sheet and rename it
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ACL Report"

    #Add the headers, set font, colour and column width (from header dictionary)
    for col, head in zip(range(1,len(header) + 1), header.items()):
        ws1['{}1'.format(get_column_letter(col))] = head[0]      # get_column_letter converts number to letter
        ws1['{}1'.format(get_column_letter(col))].fill = PatternFill(bgColor=colors.Color("00DCDCDC"))
        ws1['{}1'.format(get_column_letter(col))].font = Font(bold=True, size=14)
        ws1.column_dimensions[get_column_letter(col)].width = head[1]
    # Add the ACE entries. The columns holding numbers are changed to integrars
    for ace in final_acl:
        ace[1] = int(ace[1])
        ace[8] = int(ace[8])
        ws1.append(ace)

    # Add a key at start with info on the colourised rows for ACEs with frequent hitcnts
    ws1.insert_rows(1)
    ws1.insert_rows(2)
    keys = {'A1': 'Key:', 'B1':'Hit in last 1 day', 'E1':'Hit in last 7 days', 'G1':'Hit in last 30 days', 'I1':'Inactive'}
    colour  = {'B1':'E6B0AA', 'E1':'A9CCE3', 'G1':'F5CBA7', 'I1':'D4EFDF'}

    for cell, val in keys.items():
        ws1[cell] = val
    ws1['A1'].font = Font(bold=True)
    for cell, col in colour.items():
        ws1[cell].fill = PatternFill(start_color=col, end_color=col, fill_type='solid')

    ws1.freeze_panes = ws1['A4']                    # Freezes the top row (A1) so remains when scrolling
    ws1.auto_filter.ref = 'A3:L4'                   # Adds dropdown to headers to the headers

    # Colours used for columns dependant on the last hit data (J column). Formula is a standard XL formula
    style_red = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00E6B0AA")))
    rule_1day = Rule(type="expression",formula=["=AND(TODAY()-$J1>=0,TODAY()-$J1<=1)"], dxf=style_red)
    style_blu = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00A9CCE3")))
    rule_7day = Rule(type="expression", formula=["=AND(TODAY()-$J1>=0,TODAY()-$J1<=7)"], dxf=style_blu)
    style_org = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00F5CBA7")))
    rule_30day = Rule(type="expression", formula=["=AND(TODAY()-$J1>=0,TODAY()-$J1<=30)"], dxf=style_org)
    style_grn = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00D4EFDF")))
    rule_inactive = Rule(type="expression",formula=['=$L1="inactive"'], dxf=style_grn)

    # Apply the rules to workbook and save it
    for rule in [rule_1day, rule_7day, rule_30day, rule_inactive]:
        ws1.conditional_formatting.add(ws1.dimensions, rule)
    wb.save(filename)
    print('File {} has been created'.format(filename))

# ###################################### Run the scripts ######################################

def main():
    print('\n' + '=' * 30, 'ASA ACL Auditor v0.3 (tested 9.8)', '=' * 30)
	# 1. Gather input from user
    args = create_parser()
    # 2. Validate the input data is correct
    run = Validate(args)
    args = run.verify_args()
    # 3. Filter ACLs based on name and IP entered by user
    acl_data = run.get_filter_acl()
    # 4. Sanitise data into a correct format for the XL sheet
    run1 = Format_data(acl_data)
    acl = run1.format_acl()
    # 5. Adds timestamp for last time ACE was hit
    final_acl = run1.lasthit_time(acl)
    # 6. Create XL spreadsheet
    create_xls(args, final_acl)

if __name__ == '__main__':
    main()
